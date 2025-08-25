"""
Data Export Module for GhostCrawler

Provides functionality to export crawled data in various formats:
- CSV: Tabular data with headers
- JSON: Structured data with pretty-print option
- Excel: Multi-sheet workbooks with formatting
- HTML: Formatted reports with styling

Features:
- Nested data flattening
- Streaming for large datasets
- Compression options
- Custom field mapping
- Data transformation
- Batch export with progress tracking
"""

import csv
import json
import os
import gzip
import zipfile
import logging
import time
import re
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Callable, Set, Tuple, Iterator
from datetime import datetime
import asyncio
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from enum import Enum, auto
import html as html_lib

# Optional dependencies - imported on demand
excel_available = False
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    excel_available = True
except ImportError:
    pass

# Setup logging
logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """Supported export formats"""
    CSV = auto()
    JSON = auto()
    EXCEL = auto()
    HTML = auto()


class CompressionType(Enum):
    """Supported compression types"""
    NONE = auto()
    GZIP = auto()
    ZIP = auto()


@dataclass
class ExportOptions:
    """Configuration options for data export"""
    # Field selection
    include_fields: Optional[List[str]] = None
    exclude_fields: Optional[List[str]] = None
    
    # Field mapping (source_field -> target_field)
    field_mapping: Dict[str, str] = field(default_factory=dict)
    
    # Data transformation
    transformers: Dict[str, Callable] = field(default_factory=dict)
    
    # Formatting
    pretty_print: bool = True
    include_metadata: bool = True
    
    # Compression
    compression: CompressionType = CompressionType.NONE
    
    # Excel-specific options
    excel_sheet_name: str = "Crawled Data"
    excel_create_summary: bool = True
    
    # CSV-specific options
    csv_delimiter: str = ","
    csv_quotechar: str = '"'
    
    # HTML-specific options
    html_title: str = "GhostCrawler Export Report"
    html_include_css: bool = True
    
    # Batch processing
    batch_size: int = 100
    max_workers: int = 4
    
    # Flattening options
    flatten_nested: bool = True
    flatten_separator: str = "_"
    max_flatten_depth: int = 5


class DataExporter:
    """
    Data exporter for GhostCrawler
    
    Exports crawled data to various formats with customizable options.
    """
    
    def __init__(self, output_dir: str = "exports"):
        """
        Initialize the data exporter
        
        Args:
            output_dir: Directory for export files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._executor = ThreadPoolExecutor(max_workers=8)
        logger.info(f"DataExporter initialized with output directory: {output_dir}")
    
    def _prepare_export_path(self, filename: str, format_ext: str, 
                            compression: CompressionType) -> Path:
        """
        Prepare export file path with appropriate extension
        
        Args:
            filename: Base filename
            format_ext: Format extension (.csv, .json, etc.)
            compression: Compression type
            
        Returns:
            Path object for export file
        """
        # Clean filename
        clean_filename = re.sub(r'[^\w\-\.]', '_', filename)
        
        # Add format extension if not present
        if not clean_filename.endswith(format_ext):
            clean_filename += format_ext
        
        # Add compression extension if needed
        if compression == CompressionType.GZIP:
            clean_filename += '.gz'
        elif compression == CompressionType.ZIP:
            clean_filename += '.zip'
        
        return self.output_dir / clean_filename
    
    def _flatten_dict(self, data: Dict[str, Any], prefix: str = "", 
                     separator: str = "_", max_depth: int = 5, 
                     current_depth: int = 0) -> Dict[str, Any]:
        """
        Flatten nested dictionary
        
        Args:
            data: Nested dictionary
            prefix: Prefix for flattened keys
            separator: Separator for nested keys
            max_depth: Maximum depth to flatten
            current_depth: Current recursion depth
            
        Returns:
            Flattened dictionary
        """
        if current_depth >= max_depth:
            return {prefix: str(data)}
        
        flattened = {}
        
        for key, value in data.items():
            # Create new key with prefix
            new_key = f"{prefix}{separator}{key}" if prefix else key
            
            # Recursively flatten nested dictionaries
            if isinstance(value, dict) and value:
                nested_flat = self._flatten_dict(
                    value, new_key, separator, max_depth, current_depth + 1
                )
                flattened.update(nested_flat)
            # Handle lists - convert to string or flatten if contains dicts
            elif isinstance(value, list):
                if value and isinstance(value[0], dict):
                    # List of dicts - flatten each and add index
                    for i, item in enumerate(value):
                        if isinstance(item, dict):
                            nested_flat = self._flatten_dict(
                                item, f"{new_key}{separator}{i}", 
                                separator, max_depth, current_depth + 1
                            )
                            flattened.update(nested_flat)
                else:
                    # Regular list - join with commas
                    flattened[new_key] = ", ".join(str(x) for x in value)
            else:
                # Regular value
                flattened[new_key] = value
        
        return flattened
    
    def _apply_field_mapping(self, data: Dict[str, Any], 
                            mapping: Dict[str, str]) -> Dict[str, Any]:
        """
        Apply field mapping to rename keys
        
        Args:
            data: Data dictionary
            mapping: Mapping of source_field -> target_field
            
        Returns:
            Dictionary with mapped field names
        """
        if not mapping:
            return data
        
        result = {}
        for key, value in data.items():
            # Use mapped name if available, otherwise keep original
            new_key = mapping.get(key, key)
            result[new_key] = value
        
        return result
    
    def _apply_transformers(self, data: Dict[str, Any], 
                           transformers: Dict[str, Callable]) -> Dict[str, Any]:
        """
        Apply transformation functions to specific fields
        
        Args:
            data: Data dictionary
            transformers: Dict of field -> transformer function
            
        Returns:
            Dictionary with transformed values
        """
        if not transformers:
            return data
        
        result = data.copy()
        for field, transformer in transformers.items():
            if field in result:
                try:
                    result[field] = transformer(result[field])
                except Exception as e:
                    logger.warning(f"Transformer for field '{field}' failed: {e}")
        
        return result
    
    def _filter_fields(self, data: Dict[str, Any], 
                      include: Optional[List[str]] = None,
                      exclude: Optional[List[str]] = None) -> Dict[str, Any]:
        """
        Filter fields based on include/exclude lists
        
        Args:
            data: Data dictionary
            include: Fields to include (if None, include all)
            exclude: Fields to exclude
            
        Returns:
            Filtered dictionary
        """
        if include:
            # Only keep fields in include list
            return {k: v for k, v in data.items() if k in include}
        
        if exclude:
            # Remove fields in exclude list
            return {k: v for k, v in data.items() if k not in exclude}
        
        return data
    
    def _add_metadata(self, data: Dict[str, Any], url: str) -> Dict[str, Any]:
        """
        Add metadata to exported data
        
        Args:
            data: Original data dictionary
            url: Source URL
            
        Returns:
            Data with added metadata
        """
        metadata = {
            "export_timestamp": datetime.now().isoformat(),
            "source_url": url,
            "exporter_version": "1.0.0"
        }
        
        # Add metadata as top-level fields with meta_ prefix
        result = data.copy()
        for key, value in metadata.items():
            result[f"meta_{key}"] = value
        
        return result
    
    def _prepare_data(self, data: Dict[str, Any], url: str, 
                     options: ExportOptions) -> Dict[str, Any]:
        """
        Prepare data for export by applying all transformations
        
        Args:
            data: Raw data dictionary
            url: Source URL
            options: Export options
            
        Returns:
            Processed data ready for export
        """
        processed = data.copy()
        
        # Extract JSON-LD data if present
        if 'jsonld' in processed and isinstance(processed['jsonld'], dict):
            # Move car data to top level with jsonld_ prefix
            if 'car' in processed['jsonld']:
                car_data = processed['jsonld']['car']
                for key, value in car_data.items():
                    processed[f"jsonld_{key}"] = value
        
        # Flatten nested structures if requested
        if options.flatten_nested:
            processed = self._flatten_dict(
                processed, 
                separator=options.flatten_separator,
                max_depth=options.max_flatten_depth
            )
        
        # Apply field mapping
        processed = self._apply_field_mapping(processed, options.field_mapping)
        
        # Apply transformers
        processed = self._apply_transformers(processed, options.transformers)
        
        # Add metadata if requested
        if options.include_metadata:
            processed = self._add_metadata(processed, url)
        
        # Filter fields
        processed = self._filter_fields(
            processed, 
            include=options.include_fields,
            exclude=options.exclude_fields
        )
        
        return processed
    
    def _get_all_fields(self, data_list: List[Dict[str, Any]]) -> List[str]:
        """
        Get all unique fields from a list of data dictionaries
        
        Args:
            data_list: List of data dictionaries
            
        Returns:
            List of all unique field names
        """
        all_fields = set()
        for data in data_list:
            all_fields.update(data.keys())
        return sorted(list(all_fields))
    
    def export_to_csv(self, data: Dict[str, Any], url: str, 
                     filename: str = "export", options: Optional[ExportOptions] = None) -> str:
        """
        Export data to CSV format
        
        Args:
            data: Data dictionary
            url: Source URL
            filename: Output filename (without extension)
            options: Export options
            
        Returns:
            Path to exported file
        """
        options = options or ExportOptions()
        
        # Prepare data
        processed_data = self._prepare_data(data, url, options)
        
        # Prepare file path
        file_path = self._prepare_export_path(filename, ".csv", options.compression)
        
        try:
            # Handle compression
            if options.compression == CompressionType.GZIP:
                open_func = gzip.open
                mode = 'wt'
            else:
                open_func = open
                mode = 'w'
            
            with open_func(file_path, mode, newline='', encoding='utf-8') as f:
                # Get field names (all keys in the processed data)
                fieldnames = list(processed_data.keys())
                
                writer = csv.DictWriter(
                    f, 
                    fieldnames=fieldnames,
                    delimiter=options.csv_delimiter,
                    quotechar=options.csv_quotechar,
                    quoting=csv.QUOTE_MINIMAL
                )
                
                # Write header and data
                writer.writeheader()
                writer.writerow(processed_data)
            
            logger.info(f"Exported data to CSV: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise
    
    def export_to_json(self, data: Dict[str, Any], url: str, 
                      filename: str = "export", options: Optional[ExportOptions] = None) -> str:
        """
        Export data to JSON format
        
        Args:
            data: Data dictionary
            url: Source URL
            filename: Output filename (without extension)
            options: Export options
            
        Returns:
            Path to exported file
        """
        options = options or ExportOptions()
        
        # Prepare data
        processed_data = self._prepare_data(data, url, options)
        
        # Prepare file path
        file_path = self._prepare_export_path(filename, ".json", options.compression)
        
        try:
            # Handle compression
            if options.compression == CompressionType.GZIP:
                open_func = gzip.open
                mode = 'wt'
            else:
                open_func = open
                mode = 'w'
            
            with open_func(file_path, mode, encoding='utf-8') as f:
                # Write JSON with or without pretty printing
                if options.pretty_print:
                    json.dump(processed_data, f, indent=2, ensure_ascii=False)
                else:
                    json.dump(processed_data, f, ensure_ascii=False)
            
            logger.info(f"Exported data to JSON: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            raise
    
    def export_to_excel(self, data: Dict[str, Any], url: str, 
                       filename: str = "export", options: Optional[ExportOptions] = None) -> str:
        """
        Export data to Excel format
        
        Args:
            data: Data dictionary
            url: Source URL
            filename: Output filename (without extension)
            options: Export options
            
        Returns:
            Path to exported file
        """
        if not excel_available:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        options = options or ExportOptions()
        
        # Prepare data
        processed_data = self._prepare_data(data, url, options)
        
        # Prepare file path (compression handled differently for Excel)
        file_path = self._prepare_export_path(filename, ".xlsx", CompressionType.NONE)
        
        try:
            # Create workbook and sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = options.excel_sheet_name
            
            # Write headers
            headers = list(processed_data.keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                # Style header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Write data
            for col_idx, field in enumerate(headers, 1):
                value = processed_data.get(field, "")
                ws.cell(row=2, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                # Set minimum width based on header length
                ws.column_dimensions[column_letter].width = max(10, min(50, len(header) + 2))
            
            # Create summary sheet if requested
            if options.excel_create_summary:
                self._create_excel_summary(wb, url, processed_data)
            
            # Save workbook
            wb.save(file_path)
            
            # Handle compression if requested
            compressed_path = None
            if options.compression != CompressionType.NONE:
                compressed_path = self._compress_file(
                    file_path, 
                    options.compression
                )
                # Remove original file
                os.unlink(file_path)
                file_path = compressed_path
            
            logger.info(f"Exported data to Excel: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _create_excel_summary(self, workbook: 'openpyxl.Workbook', 
                             url: str, data: Dict[str, Any]) -> None:
        """
        Create a summary sheet in Excel workbook
        
        Args:
            workbook: Excel workbook
            url: Source URL
            data: Processed data
        """
        ws = workbook.create_sheet(title="Summary")
        
        # Add title
        ws['A1'] = "GhostCrawler Export Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        # Add metadata
        row = 3
        ws.cell(row=row, column=1, value="Export Date:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        row += 1
        ws.cell(row=row, column=1, value="Source URL:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=url)
        
        # Add data summary
        row += 2
        ws.cell(row=row, column=1, value="Data Summary").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws.cell(row=row, column=1, value="Field").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
        
        # Add important fields from data
        important_fields = [
            "meta_export_timestamp",
            "jsonld_name", 
            "jsonld_brand",
            "jsonld_model",
            "jsonld_vehicleEngine",
            "jsonld_fuelType"
        ]
        
        for field in important_fields:
            if field in data:
                row += 1
                # Clean up field name for display
                display_name = field.replace("jsonld_", "").replace("meta_", "")
                display_name = display_name.replace("_", " ").title()
                
                ws.cell(row=row, column=1, value=display_name)
                ws.cell(row=row, column=2, value=str(data[field]))
        
        # Auto-adjust column widths
        for col in range(1, 3):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 25
    
    def _compress_file(self, file_path: Union[str, Path], 
                      compression: CompressionType) -> Path:
        """
        Compress a file using the specified compression type
        
        Args:
            file_path: Path to file to compress
            compression: Compression type
            
        Returns:
            Path to compressed file
        """
        file_path = Path(file_path)
        
        if compression == CompressionType.GZIP:
            compressed_path = file_path.with_suffix(file_path.suffix + '.gz')
            with open(file_path, 'rb') as f_in:
                with gzip.open(compressed_path, 'wb') as f_out:
                    f_out.write(f_in.read())
            return compressed_path
            
        elif compression == CompressionType.ZIP:
            compressed_path = file_path.with_suffix(file_path.suffix + '.zip')
            with zipfile.ZipFile(compressed_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                zip_file.write(file_path, arcname=file_path.name)
            return compressed_path
            
        return file_path
    
    def export_to_html(self, data: Dict[str, Any], url: str, 
                      filename: str = "export", options: Optional[ExportOptions] = None) -> str:
        """
        Export data to HTML format
        
        Args:
            data: Data dictionary
            url: Source URL
            filename: Output filename (without extension)
            options: Export options
            
        Returns:
            Path to exported file
        """
        options = options or ExportOptions()
        
        # Prepare data
        processed_data = self._prepare_data(data, url, options)
        
        # Prepare file path
        file_path = self._prepare_export_path(filename, ".html", options.compression)
        
        try:
            # Generate HTML content
            html_content = self._generate_html_report(processed_data, url, options)
            
            # Handle compression
            if options.compression == CompressionType.GZIP:
                open_func = gzip.open
                mode = 'wt'
            else:
                open_func = open
                mode = 'w'
            
            with open_func(file_path, mode, encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"Exported data to HTML: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error exporting to HTML: {e}")
            raise
    
    def _generate_html_report(self, data: Dict[str, Any], url: str, 
                             options: ExportOptions) -> str:
        """
        Generate HTML report content
        
        Args:
            data: Processed data dictionary
            url: Source URL
            options: Export options
            
        Returns:
            HTML content as string
        """
        # Define CSS styles
        css = """
        body {
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
        }
        h1 {
            color: #2c3e50;
            border-bottom: 2px solid #3498db;
            padding-bottom: 10px;
        }
        h2 {
            color: #2980b9;
            margin-top: 30px;
        }
        .metadata {
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 20px;
        }
        .metadata p {
            margin: 5px 0;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th {
            background-color: #3498db;
            color: white;
            text-align: left;
            padding: 10px;
        }
        td {
            border: 1px solid #ddd;
            padding: 10px;
        }
        tr:nth-child(even) {
            background-color: #f2f2f2;
        }
        .section {
            margin-bottom: 30px;
            border: 1px solid #eee;
            padding: 15px;
            border-radius: 5px;
        }
        .field-name {
            font-weight: bold;
            color: #2c3e50;
        }
        .timestamp {
            color: #7f8c8d;
            font-size: 0.9em;
        }
        """
        
        # Start building HTML
        html = f"""<!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{html_lib.escape(options.html_title)}</title>
            {f"<style>{css}</style>" if options.html_include_css else ""}
        </head>
        <body>
            <h1>{html_lib.escape(options.html_title)}</h1>
            
            <div class="metadata">
                <h2>Export Information</h2>
                <p><span class="field-name">URL:</span> {html_lib.escape(url)}</p>
                <p><span class="field-name">Export Date:</span> <span class="timestamp">{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</span></p>
            </div>
        """
        
        # Group data by categories for better presentation
        categories = {
            "Vehicle Information": [],
            "Technical Specifications": [],
            "Metadata": [],
            "Other": []
        }
        
        for key, value in data.items():
            if key.startswith("meta_"):
                categories["Metadata"].append((key, value))
            elif key.startswith("jsonld_"):
                categories["Vehicle Information"].append((key, value))
            elif "specs" in key.lower() or "engine" in key.lower() or "dimension" in key.lower():
                categories["Technical Specifications"].append((key, value))
            else:
                categories["Other"].append((key, value))
        
        # Add each category section
        for category, fields in categories.items():
            if not fields:
                continue
                
            html += f"""
            <div class="section">
                <h2>{html_lib.escape(category)}</h2>
                <table>
                    <tr>
                        <th>Field</th>
                        <th>Value</th>
                    </tr>
            """
            
            for key, value in fields:
                # Format field name for display
                display_name = key
                if key.startswith("jsonld_"):
                    display_name = key[7:]  # Remove jsonld_ prefix
                elif key.startswith("meta_"):
                    display_name = key[5:]  # Remove meta_ prefix
                
                display_name = display_name.replace("_", " ").title()
                
                # Format value
                if isinstance(value, (dict, list)):
                    display_value = f"<pre>{html_lib.escape(json.dumps(value, indent=2))}</pre>"
                else:
                    display_value = html_lib.escape(str(value))
                
                html += f"""
                <tr>
                    <td class="field-name">{html_lib.escape(display_name)}</td>
                    <td>{display_value}</td>
                </tr>
                """
            
            html += """
                </table>
            </div>
            """
        
        # Close HTML
        html += """
        </body>
        </html>
        """
        
        return html
    
    async def export_batch(self, data_dict: Dict[str, Dict[str, Any]], 
                          format_type: ExportFormat, output_dir: Optional[str] = None,
                          options: Optional[ExportOptions] = None) -> Dict[str, str]:
        """
        Export multiple items in batch
        
        Args:
            data_dict: Dictionary mapping URLs to data dictionaries
            format_type: Export format
            output_dir: Output directory (overrides default)
            options: Export options
            
        Returns:
            Dictionary mapping URLs to export file paths
        """
        options = options or ExportOptions()
        
        if output_dir:
            original_output_dir = self.output_dir
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Prepare batch processing
        batch_size = options.batch_size
        urls = list(data_dict.keys())
        results = {}
        
        # Export method mapping
        export_methods = {
            ExportFormat.CSV: self.export_to_csv,
            ExportFormat.JSON: self.export_to_json,
            ExportFormat.EXCEL: self.export_to_excel,
            ExportFormat.HTML: self.export_to_html
        }
        
        export_method = export_methods[format_type]
        
        # Process in batches
        for i in range(0, len(urls), batch_size):
            batch_urls = urls[i:i + batch_size]
            batch_tasks = []
            
            for url in batch_urls:
                # Generate filename from URL
                parsed_url = url.split('/')
                filename = f"export_{parsed_url[-1]}" if len(parsed_url) > 3 else f"export_{hash(url)}"
                
                # Create task for this URL
                task = asyncio.create_task(
                    self._export_url_safe(
                        export_method, 
                        data_dict[url], 
                        url, 
                        filename, 
                        options
                    )
                )
                batch_tasks.append((url, task))
            
            # Wait for batch to complete
            for url, task in batch_tasks:
                try:
                    file_path = await task
                    if file_path:
                        results[url] = file_path
                except Exception as e:
                    logger.error(f"Error exporting {url}: {e}")
            
            # Progress update
            logger.info(f"Batch export progress: {len(results)}/{len(urls)} completed")
        
        # Restore original output directory if changed
        if output_dir:
            self.output_dir = original_output_dir
        
        return results
    
    async def _export_url_safe(self, export_func, data, url, filename, options):
        """Wrapper to safely execute export function with error handling"""
        try:
            return export_func(data, url, filename, options)
        except Exception as e:
            logger.error(f"Export failed for {url}: {e}")
            return None
    
    def create_combined_export(self, data_dict: Dict[str, Dict[str, Any]], 
                              format_type: ExportFormat, filename: str = "combined_export",
                              options: Optional[ExportOptions] = None) -> str:
        """
        Create a single export file combining multiple items
        
        Args:
            data_dict: Dictionary mapping URLs to data dictionaries
            format_type: Export format
            filename: Output filename
            options: Export options
            
        Returns:
            Path to export file
        """
        options = options or ExportOptions()
        
        if format_type == ExportFormat.CSV:
            return self._create_combined_csv(data_dict, filename, options)
        elif format_type == ExportFormat.JSON:
            return self._create_combined_json(data_dict, filename, options)
        elif format_type == ExportFormat.EXCEL:
            return self._create_combined_excel(data_dict, filename, options)
        elif format_type == ExportFormat.HTML:
            return self._create_combined_html(data_dict, filename, options)
        else:
            raise ValueError(f"Unsupported format for combined export: {format_type}")
    
    def _create_combined_csv(self, data_dict: Dict[str, Dict[str, Any]], 
                            filename: str, options: ExportOptions) -> str:
        """Create combined CSV export"""
        # Prepare file path
        file_path = self._prepare_export_path(filename, ".csv", options.compression)
        
        try:
            # Process all data items
            processed_items = []
            for url, data in data_dict.items():
                processed = self._prepare_data(data, url, options)
                processed_items.append(processed)
            
            # Get all unique fields
            all_fields = self._get_all_fields(processed_items)
            
            # Handle compression
            if options.compression == CompressionType.GZIP:
                open_func = gzip.open
                mode = 'wt'
            else:
                open_func = open
                mode = 'w'
            
            with open_func(file_path, mode, newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(
                    f, 
                    fieldnames=all_fields,
                    delimiter=options.csv_delimiter,
                    quotechar=options.csv_quotechar,
                    quoting=csv.QUOTE_MINIMAL
                )
                
                # Write header and data
                writer.writeheader()
                for item in processed_items:
                    writer.writerow(item)
            
            logger.info(f"Exported combined data to CSV: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error creating combined CSV export: {e}")
            raise
    
    def _create_combined_json(self, data_dict: Dict[str, Dict[str, Any]], 
                             filename: str, options: ExportOptions) -> str:
        """Create combined JSON export"""
        # Prepare file path
        file_path = self._prepare_export_path(filename, ".json", options.compression)
        
        try:
            # Process all data items
            combined_data = {}
            for url, data in data_dict.items():
                processed = self._prepare_data(data, url, options)
                # Use URL as key in combined data
                combined_data[url] = processed
            
            # Handle compression
            if options.compression == CompressionType.GZIP:
                open_func = gzip.open
                mode = 'wt'
            else:
                open_func = open
                mode = 'w'
            
            with open_func(file_path, mode, encoding='utf-8') as f:
                # Write JSON with or without pretty printing
                if options.pretty_print:
                    json.dump(combined_data, f, indent=2, ensure_ascii=False)
                else:
                    json.dump(combined_data, f, ensure_ascii=False)
            
            logger.info(f"Exported combined data to JSON: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error creating combined JSON export: {e}")
            raise
    
    def _create_combined_excel(self, data_dict: Dict[str, Dict[str, Any]], 
                              filename: str, options: ExportOptions) -> str:
        """Create combined Excel export"""
        if not excel_available:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        # Prepare file path
        file_path = self._prepare_export_path(filename, ".xlsx", CompressionType.NONE)
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Create summary sheet
            summary_sheet = wb.active
            summary_sheet.title = "Summary"
            
            # Add title
            summary_sheet['A1'] = "GhostCrawler Combined Export"
            summary_sheet['A1'].font = Font(bold=True, size=14)
            summary_sheet.merge_cells('A1:D1')
            
            # Add metadata
            row = 3
            summary_sheet.cell(row=row, column=1, value="Export Date:").font = Font(bold=True)
            summary_sheet.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            row += 1
            summary_sheet.cell(row=row, column=1, value="Total URLs:").font = Font(bold=True)
            summary_sheet.cell(row=row, column=2, value=len(data_dict))
            
            # Add URL index
            row += 2
            summary_sheet.cell(row=row, column=1, value="URL Index").font = Font(bold=True, size=12)
            summary_sheet.merge_cells(f'A{row}:D{row}')
            
            row += 1
            summary_sheet.cell(row=row, column=1, value="#").font = Font(bold=True)
            summary_sheet.cell(row=row, column=2, value="URL").font = Font(bold=True)
            summary_sheet.cell(row=row, column=3, value="Sheet Name").font = Font(bold=True)
            
            # Process each URL and create a sheet
            for idx, (url, data) in enumerate(data_dict.items(), 1):
                # Add to index
                row += 1
                summary_sheet.cell(row=row, column=1, value=idx)
                summary_sheet.cell(row=row, column=2, value=url)
                
                # Create sheet name from URL (sanitized)
                sheet_name = f"URL_{idx}"
                summary_sheet.cell(row=row, column=3, value=sheet_name)
                
                # Create data sheet
                self._add_data_sheet(wb, sheet_name, data, url, options)
            
            # Auto-adjust column widths in summary
            for col in range(1, 4):
                if col == 2:  # URL column
                    summary_sheet.column_dimensions[get_column_letter(col)].width = 50
                else:
                    summary_sheet.column_dimensions[get_column_letter(col)].width = 15
            
            # Save workbook
            wb.save(file_path)
            
            # Handle compression if requested
            if options.compression != CompressionType.NONE:
                compressed_path = self._compress_file(
                    file_path, 
                    options.compression
                )
                # Remove original file
                os.unlink(file_path)
                file_path = compressed_path
            
            logger.info(f"Exported combined data to Excel: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error creating combined Excel export: {e}")
            raise
    
    def _add_data_sheet(self, workbook: 'openpyxl.Workbook', sheet_name: str, 
                       data: Dict[str, Any], url: str, options: ExportOptions) -> None:
        """Add a data sheet to Excel workbook for a single URL"""
        # Process data
        processed_data = self._prepare_data(data, url, options)
        
        # Create sheet
        ws = workbook.create_sheet(title=sheet_name)
        
        # Add URL as title
        ws['A1'] = f"Data for: {url}"
        ws['A1'].font = Font(bold=True)
        ws.merge_cells('A1:D1')
        
        # Write headers and data
        headers = list(processed_data.keys())
        
        # Write in two columns: Field | Value
        ws.cell(row=3, column=1, value="Field").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Value").font = Font(bold=True)
        
        for idx, field in enumerate(headers):
            row = idx + 4
            
            # Field name
            display_name = field
            if field.startswith("jsonld_"):
                display_name = field[7:]  # Remove jsonld_ prefix
            elif field.startswith("meta_"):
                display_name = field[5:]  # Remove meta_ prefix
            
            display_name = display_name.replace("_", " ").title()
            
            ws.cell(row=row, column=1, value=display_name)
            
            # Field value
            value = processed_data.get(field, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value)
            ws.cell(row=row, column=2, value=str(value))
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 70
    
    def _create_combined_html(self, data_dict: Dict[str, Dict[str, Any]], 
                             filename: str, options: ExportOptions) -> str:
        """Create combined HTML export"""
        # Prepare file path
        file_path = self._prepare_export_path(filename, ".html", options.compression)
        
        try:
            # Define CSS styles
            css = """
            body {
                font-family: Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 1200px;
                margin: 0 auto;
                padding: 20px;
            }
            h1, h2, h3 {
                color: #2c3e50;
            }
            h1 {
                border-bottom: 2px solid #3498db;
                padding-bottom: 10px;
            }
            h2 {
                margin-top: 30px;
                border-bottom: 1px solid #ddd;
                padding-bottom: 5px;
            }
            .metadata {
                background-color: #f8f9fa;
                padding: 15px;
                border-radius: 5px;
                margin-bottom: 20px;
            }
            .url-section {
                margin-bottom: 40px;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 20px;
            }
            .url-header {
                background-color: #f8f9fa;
                padding: 10px;
                margin: -20px -20px 20px -20px;
                border-bottom: 1px solid #ddd;
                border-radius: 5px 5px 0 0;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }
            th {
                background-color: #3498db;
                color: white;
                text-align: left;
                padding: 10px;
            }
            td {
                border: 1px solid #ddd;
                padding: 10px;
            }
            tr:nth-child(even) {
                background-color: #f2f2f2;
            }
            .index-table {
                width: 100%;
                margin-bottom: 30px;
            }
            .timestamp {
                color: #7f8c8d;
                font-size: 0.9em;
            }
            """
            
            # Start building HTML
            html = f"""<!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>{html_lib.escape(options.html_title)} - Combined Report</title>
                {f"<style>{css}</style>" if options.html_include_css else ""}
            </head>
            <body>
                <h1>{html_lib.escape(options.html_title)} - Combined Report</h1>
                
                <div class="metadata">
                    <h2>Export Information</h2>
                    <p><strong>Export Date:</strong> <span class="timestamp">{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</span></p>
                    <p><strong>Total URLs:</strong> {len(data_dict)}</p>
                </div>
                
                <h2>URL Index</h2>
                <table class="index-table">
                    <tr>
                        <th>#</th>
                        <th>URL</th>
                        <th>Jump to</th>
                    </tr>
            """
            
            # Add URL index
            for idx, url in enumerate(data_dict.keys(), 1):
                anchor = f"url_{idx}"
                html += f"""
                <tr>
                    <td>{idx}</td>
                    <td>{html_lib.escape(url)}</td>
                    <td><a href="#{anchor}">View Data</a></td>
                </tr>
                """
            
            html += "</table>"
            
            # Add each URL's data
            for idx, (url, data) in enumerate(data_dict.items(), 1):
                anchor = f"url_{idx}"
                processed_data = self._prepare_data(data, url, options)
                
                html += f"""
                <div id="{anchor}" class="url-section">
                    <div class="url-header">
                        <h2>URL #{idx}: {html_lib.escape(url)}</h2>
                    </div>
                """
                
                # Group data by categories for better presentation
                categories = {
                    "Vehicle Information": [],
                    "Technical Specifications": [],
                    "Metadata": [],
                    "Other": []
                }
                
                for key, value in processed_data.items():
                    if key.startswith("meta_"):
                        categories["Metadata"].append((key, value))
                    elif key.startswith("jsonld_"):
                        categories["Vehicle Information"].append((key, value))
                    elif "specs" in key.lower() or "engine" in key.lower() or "dimension" in key.lower():
                        categories["Technical Specifications"].append((key, value))
                    else:
                        categories["Other"].append((key, value))
                
                # Add each category section
                for category, fields in categories.items():
                    if not fields:
                        continue
                        
                    html += f"""
                    <h3>{html_lib.escape(category)}</h3>
                    <table>
                        <tr>
                            <th>Field</th>
                            <th>Value</th>
                        </tr>
                    """
                    
                    for key, value in fields:
                        # Format field name for display
                        display_name = key
                        if key.startswith("jsonld_"):
                            display_name = key[7:]  # Remove jsonld_ prefix
                        elif key.startswith("meta_"):
                            display_name = key[5:]  # Remove meta_ prefix
                        
                        display_name = display_name.replace("_", " ").title()
                        
                        # Format value
                        if isinstance(value, (dict, list)):
                            display_value = f"<pre>{html_lib.escape(json.dumps(value, indent=2))}</pre>"
                        else:
                            display_value = html_lib.escape(str(value))
                        
                        html += f"""
                        <tr>
                            <td><strong>{html_lib.escape(display_name)}</strong></td>
                            <td>{display_value}</td>
                        </tr>
                        """
                    
                    html += "</table>"
                
                html += "</div>"
            
            # Close HTML
            html += """
            </body>
            </html>
            """
            
            # Handle compression
            if options.compression == CompressionType.GZIP:
                open_func = gzip.open
                mode = 'wt'
            else:
                open_func = open
                mode = 'w'
            
            with open_func(file_path, mode, encoding='utf-8') as f:
                f.write(html)
            
            logger.info(f"Exported combined data to HTML: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Error creating combined HTML export: {e}")
            raise


# Common data transformers
class DataTransformers:
    """Common data transformation functions for export"""
    
    @staticmethod
    def normalize_price(value: Any) -> str:
        """Normalize price values to standard format"""
        if not value:
            return ""
        
        # Convert to string and strip non-numeric chars except decimal point
        value_str = str(value)
        digits = ''.join(c for c in value_str if c.isdigit() or c == '.')
        
        try:
            # Parse as float and format with commas
            price = float(digits)
            return f"${price:,.2f}"
        except (ValueError, TypeError):
            return value_str
    
    @staticmethod
    def format_date(value: Any, format_str: str = "%Y-%m-%d") -> str:
        """Format date strings to consistent format"""
        if not value:
            return ""
        
        value_str = str(value)
        
        # Try common date formats
        formats = [
            "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y",
            "%b %d, %Y", "%B %d, %Y", "%d %b %Y", "%d %B %Y"
        ]
        
        for fmt in formats:
            try:
                date_obj = datetime.strptime(value_str, fmt)
                return date_obj.strftime(format_str)
            except ValueError:
                continue
        
        return value_str
    
    @staticmethod
    def clean_html(value: Any) -> str:
        """Remove HTML tags from text"""
        if not value:
            return ""
        
        value_str = str(value)
        
        # Simple regex to remove HTML tags
        clean_text = re.sub(r'<[^>]+>', '', value_str)
        
        # Replace common HTML entities
        entities = {
            '&nbsp;': ' ', '&amp;': '&', '&lt;': '<', '&gt;': '>',
            '&quot;': '"', '&apos;': "'", '&#39;': "'"
        }
        for entity, replacement in entities.items():
            clean_text = clean_text.replace(entity, replacement)
        
        # Normalize whitespace
        clean_text = re.sub(r'\s+', ' ', clean_text).strip()
        
        return clean_text
    
    @staticmethod
    def normalize_boolean(value: Any) -> str:
        """Convert various boolean representations to Yes/No"""
        if isinstance(value, bool):
            return "Yes" if value else "No"
        
        if not value:
            return "No"
        
        value_str = str(value).lower()
        
        if value_str in ('true', 'yes', 'y', '1', 'on'):
            return "Yes"
        elif value_str in ('false', 'no', 'n', '0', 'off'):
            return "No"
        
        return str(value)
    
    @staticmethod
    def extract_number(value: Any) -> Optional[float]:
        """Extract numeric value from string with units"""
        if not value:
            return None
        
        value_str = str(value)
        
        # Find all numbers in the string
        matches = re.findall(r'[-+]?\d*\.\d+|\d+', value_str)
        
        if matches:
            try:
                return float(matches[0])
            except (ValueError, TypeError):
                pass
        
        return None
